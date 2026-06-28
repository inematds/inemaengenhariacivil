"""Testes das exportações (xlsx via openpyxl, pdf via reportlab) do pacote de cálculo."""

from openpyxl import load_workbook
from pytest import approx

from lib.reporting.excel_export import export_bundle_to_xlsx
from lib.reporting.pdf_export import export_memorial_to_pdf
from lib.service import solve_rectangular_beam


def _bundle() -> dict:
    return solve_rectangular_beam(
        b_cm=25.0, h_cm=50.0, fck_mpa=30.0, fyk_mpa=500.0,
        vao_m=5.0, g_knm=10.0, q_knm=10.0,
    )


def test_export_bundle_to_xlsx_roundtrip(tmp_path):
    bundle = _bundle()
    caminho = str(tmp_path / "memorial.xlsx")

    retorno = export_bundle_to_xlsx(bundle, caminho)

    assert retorno == caminho
    arquivo = tmp_path / "memorial.xlsx"
    assert arquivo.exists()
    assert arquivo.stat().st_size > 0

    # reabre e confere um valor conhecido (As adotada do resultado)
    wb = load_workbook(caminho)
    ws = wb.active
    dados = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True)
             if row[0] is not None}
    assert dados["as_adopted_cm2"] == approx(bundle["resultado"]["as_adopted_cm2"])
    assert dados["aprovado"] == "APROVADO"


def test_export_memorial_to_pdf_creates_file(tmp_path):
    bundle = _bundle()
    caminho = str(tmp_path / "memorial.pdf")

    retorno = export_memorial_to_pdf(bundle["memorial_markdown"], caminho)

    assert retorno == caminho
    arquivo = tmp_path / "memorial.pdf"
    assert arquivo.exists()
    assert arquivo.stat().st_size > 0
    assert arquivo.read_bytes().startswith(b"%PDF")
